import openpyxl
import re

# 打开 old 和 new 两个 Excel 文件
old_wb = openpyxl.load_workbook('old.xlsx')
old_ws = old_wb.active
new_wb = openpyxl.load_workbook('new.xlsx')
new_ws = new_wb.active

def clean_journal_name(journal_name):
    """
    将期刊名字中的标点符号和空格去除
    
    参数:
    journal_name (str) - 期刊名字
    
    返回:
    str - 去除标点符号和空格后的期刊名字
    """
    # 定义一个正则表达式,匹配所有标点符号和空格
    pattern = r'[!"#$%&\'()*+,-./:;<=>?@\[\\\]^_`、，·{|}~ ]'
    
    # 使用re.sub()函数去除期刊名字中的标点符号和空格
    cleaned_name = re.sub(pattern, '', journal_name)
    
    return cleaned_name

# 创建一个字典,存储 old 表中的期刊名和网址
old_journal_dict = {}
for row in range(3, old_ws.max_row + 1):
    journal_name = old_ws.cell(row=row, column=1).value
    journal_name=clean_journal_name(str(journal_name))
    journal_url = old_ws.cell(row=row, column=2).value
    if journal_url!=None:
        old_journal_dict[journal_name] = journal_url

# 遍历 new 表,如果期刊名在 old 表中,就将网址添加到 new 表对应的单元格中
for row in range(3, new_ws.max_row + 1):
    journal_name = new_ws.cell(row=row, column=3).value
    journal_name=clean_journal_name(journal_name)
    print(journal_name)
    if journal_name in old_journal_dict:
        new_ws.cell(row=row, column=5).value = old_journal_dict[journal_name]

    break
# for row in range(3, new_ws.max_row + 1):
#     journal_name = new_ws.cell(row=row, column=3).value
#     cleaned_journal_name = clean_journal_name(str(journal_name))
    
#     # 遍历 old_journal_dict 的键,看是否在 cleaned_journal_name 中
#     for old_journal_name in old_journal_dict:
#         if cleaned_journal_name.find(old_journal_name) != -1 or old_journal_name.find(cleaned_journal_name) != -1:
#             new_ws.cell(row=row, column=5).value = old_journal_dict[old_journal_name]
#             break  # 找到一个匹配的就退出内层循环

# 保存更新后的 new 表
# new_wb.save('new_updated.xlsx')


